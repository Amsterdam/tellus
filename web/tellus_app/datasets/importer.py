import logging

import openpyxl

from datasets.tellus_data.models import Locatie, LengteCategorie, SnelheidsCategorie

log = logging.getLogger(__name__)

OBJSTORE_METADATA = 'meta'


class MetaImporter(object):
    def __init__(self):
        self.sheets = {}
        self._import_meta()

    def _import_meta(self):
        """
        Function that retrieves the meta data from a file (use Objectstore later on)
        """
        wb = openpyxl.load_workbook('/tmp/tellus/AMS365_Codeboek.xlsx')
        self.sheets = {sheet_name: wb.get_sheet_by_name(sheet_name) for sheet_name in wb.get_sheet_names()}

    def process_locaties(self, sheet_name='Locaties', title_row=2, first_col=2):
        for row in self.sheets[sheet_name].iter_rows(min_row=title_row + 1, min_col=first_col):
            res = [cell.value for cell in row]

            db_row, created = Locatie.objects.update_or_create(
                meetlocatie=res[0], richtingcode=res[1], telpunt=res[2],
                straat=res[3], windrichting=res[4], zijstraat1=res[5])
            if created:
                log.info("Created {}".format(str(db_row)))
            else:
                log.info("Updated {}".format(str(db_row)))

    def process_lengte_categorie(self, sheet_name='Lengtecategorieen', title_row=1, first_col=2):
        ids = list(
            self.sheets[sheet_name].iter_rows(
                min_row=title_row, max_row=title_row, min_col=first_col))[0]
        values = list(
            self.sheets[sheet_name].iter_rows(
                min_row=title_row + 1, max_row=title_row + 1, min_col=first_col))[0]
        for r in range(5):
            val = values[r].value.split('-')
            print("`{}`".format(ids[r].value))
            db_row, created = LengteCategorie.objects.update_or_create(
                categorie=ids[r].value,
                lengte_van=val[0].replace(' ', ''),
                lengte_tot=val[1].replace(' ', '').replace('m', ' m'))
            if created:
                log.info("Created {}".format(str(db_row)))
            else:
                log.info("Updated {}".format(str(db_row)))

    def process_snelheids_categorie(self, sheet_name='Snelheidscategorieen', title_row=1, first_col=1):
        for row in self.sheets[sheet_name].iter_rows(min_row=title_row + 1, max_row=5, min_col=first_col):
            res = [cell.value for cell in row]

            db_row, created = SnelheidsCategorie.objects.update_or_create(
                categorie=res[0],
                s1=res[1], s2=res[2], s3=res[3], s4=res[4], s5=res[5],
                s6=res[6], s7=res[7], s8=res[8], s9=res[9], s10=res[10])
            if created:
                log.info("Created {}".format(str(db_row)))
            else:
                log.info("Updated {}".format(str(db_row)))
