"""Import tellus csv's."""
import csv
import logging
import os

import django
import openpyxl


django.setup()

from importer_lib.speed_processor import create_speed_intervals, create_speed_category  # noqa
from importer_lib.length_processor import create_length_intervals  # noqa
from importer_lib.telling_processor import process_telling_sheet  # noqa

from django.contrib.gis.geos import Point  # noqa

from datasets.tellus_data.models import Telling  # noqa
from datasets.tellus_data.models import MeetraaiCategorie  # noqa
from datasets.tellus_data.models import RepresentatiefCategorie  # noqa
from datasets.tellus_data.models import ValidatieCategorie  # noqa
from datasets.tellus_data.models import Meetlocatie  # noqa
from datasets.tellus_data.models import Tellus  # noqa
from datasets.tellus_data.models import TelRichting  # noqa

from objectstore.objectstore import fetch_meta_data, fetch_tellus_data_file_names  # noqa
from objectstore.objectstore import fetch_tellus_data_file_object  # noqa

log = logging.getLogger(__name__)


def process_tel_richting(tellus, richting, naam, zijstraat):
    tel_richting, _ = TelRichting.objects.update_or_create(
        tellus=tellus,
        richting=richting,
        naam=naam,
        zijstraat=zijstraat
    )


class TellusImporter(object):

    def __init__(self, codebook, codebook_addon):
        self.codebook_sheets = self._import_meta(codebook)
        self.codebook_addon_sheets = self._import_meta(codebook_addon)

    def _import_meta(self, codebook_filename):
        """
        Function that retrieves the meta data from a Objectstore
        """
        os.makedirs("/tmp/tellus", exist_ok=True)
        with open("/tmp/tellus/{}".format(codebook_filename), 'wb') as f:
            f.write(fetch_meta_data(codebook_filename))

        wb = openpyxl.load_workbook("/tmp/tellus/{}".format(codebook_filename))
        return {sheet_name: wb[sheet_name] for
                sheet_name in wb.sheetnames}

    def process_tellus_locaties(self, title_row=1,
                                first_col=0):
        """
        Import the data for the Tellus (Locations)
        :param sheet_name:
        :param title_row:
        :param first_col:
        :return:
        """
        for row in self.codebook_addon_sheets['Locaties'].iter_rows(
                min_row=title_row + 1,
                min_col=first_col):
            res = [cell.value for cell in row]

            if not res[0]:
                # ignore empty rows
                continue

            # Meetlocatie
            meetlocatie, _ = Meetlocatie.objects.update_or_create(
                id=res[12],
                name=res[2]
            )

            # Tellus
            objnr_vor = res[0]  # e.g.: TP0001
            tellus_id = int(objnr_vor[2:])  # e.g.: 1

            tellus, _ = Tellus.objects.update_or_create(
                id=tellus_id,
                objnr_vor=objnr_vor,
                objnr_leverancier=res[1],
                snelheids_categorie=res[11],
                latitude=res[9],
                longitude=res[10],
                rijksdriehoek_x=res[7],
                rijksdriehoek_y=res[8],
                geometrie=Point(float(res[7]), float(res[8]), srid=28992),
                meetlocatie=meetlocatie
            )

            # TelRichting
            if res[5] != 'n.v.t.':  # Richting 1
                process_tel_richting(tellus, 1, res[5], res[3])
            if res[6] != 'n.v.t.':  # Richting 2
                process_tel_richting(tellus, 2, res[6], res[4])

            log.debug("Processed {}".format(str(res[0])))

    def process_snelheids_categorie(self, sheet_name='Snelheidscategorieen', title_row=1, first_col=1):
        """
        Import SnelheidsCategorie & SnelheidsInterval objects
        :param sheet_name:
        :param title_row:
        :param first_col:
        :return:
        """
        for row in self.codebook_sheets[sheet_name].iter_rows(
                min_row=title_row + 1,
                max_row=5,
                min_col=first_col):
            res = [cell.value for cell in row]  # [1, '< 30 km/u', '31 - 40 km/u', ..., '> 100 km/u']
            categorie = int(res[0])             # 1
            values = res[1:]                    # ['< 30 km/u', '31 - 40 km/u', ..., '> 100 km/u']

            # Identify all unique intervals:
            create_speed_intervals(values)

            # Store reference from "Snelheid categorie" to interval.
            create_speed_category(categorie, values)

    def process_meetraai_categorie(self):
        for row in self.codebook_sheets['Meetraai'].iter_rows(
                min_row=1, max_row=3, min_col=1):
            res = [cell.value for cell in row]
            db_row, created = MeetraaiCategorie.objects.update_or_create(
                id=res[0],
                label=res[1]
            )
            if created:
                log.info("Created {}".format(str(db_row)))
            else:
                log.info("Updated {}".format(str(db_row)))

    def process_representatief_categorie(self):
        for row in self.codebook_sheets['Representatief'].iter_rows(
                min_row=1, max_row=6, min_col=1):
            res = [cell.value for cell in row]
            db_row, created = RepresentatiefCategorie.objects.update_or_create(
                id=res[0],
                label=res[1])
            if created:
                log.info("Created {}".format(str(db_row)))
            else:
                log.info("Updated {}".format(str(db_row)))

    def process_validatie_categorie(self):
        for row in self.codebook_sheets['Validatie'].iter_rows(
                min_row=1, max_row=6, min_col=1):
            res = [cell.value for cell in row]
            db_row, created = ValidatieCategorie.objects.update_or_create(
                id=res[0],
                label=res[1])
            if created:
                log.info("Created {}".format(str(db_row)))
            else:
                log.info("Updated {}".format(str(db_row)))

    def process_lengte_interval(self, sheet_name='Lengtecategorieen', title_row=1, first_col=1):
        """
        Import LengteInterval objects
        :param sheet_name:
        :param title_row:
        :param first_col:
        :return:
        """
        for row in self.codebook_sheets[sheet_name].iter_rows(
                min_row=title_row + 1,
                max_row=2,
                min_col=first_col):

            res = [cell.value for cell in row]  # [1, '0 - 5,1 m', ..., '> 12,2 m']
            values = res[1:]  # ['0 - 5,1 m', ..., '> 12,2 m']

            create_length_intervals(values)

    def download_tellus_data(self, obj_store_path):
        filename = os.path.basename(obj_store_path)
        target_path = os.path.join('/tmp/tellus/', filename)
        exists = os.path.isfile(target_path)
        if exists:
            log.info("File exists, skipping download: {}".format(obj_store_path))
        else:
            with open(target_path, 'wb') as f:
                f.write(fetch_tellus_data_file_object(obj_store_path))
        return target_path


def get_importer():
    assert os.getenv('TELLUS_OBJECTSTORE_PASSWORD')
    os.makedirs('/tmp/tellus', exist_ok=True)
    return TellusImporter(
        codebook='AMS365_codeboek_v5.xlsx',
        codebook_addon='AMS365_codeboek_v8_aanvulling.xlsx'
    )


def import_core():
    importer = get_importer()

    importer.process_meetraai_categorie()
    importer.process_representatief_categorie()
    importer.process_validatie_categorie()
    importer.process_lengte_interval()
    importer.process_snelheids_categorie()
    importer.process_tellus_locaties()

    log.info("Done importing tellus core")


def import_telling(csv_path):
    file_name = os.path.basename(csv_path)
    with open(csv_path) as csv_file:
        csvReader = csv.reader(csv_file, dialect='excel', delimiter=';')
        process_telling_sheet(file_name, csvReader)


def prepare_import_tellingen():
    """
    Delete existing Tellingen objects AND get source data (csv files)
    :return:
    """
    assert os.getenv('TELLUS_OBJECTSTORE_PASSWORD')
    importer = get_importer()

    log.debug("Delete all telling objects: ")
    Telling.objects.all().delete()
    log.debug("Delete telling objects done")

    file_names = fetch_tellus_data_file_names()

    file_paths = [importer.download_tellus_data(file_name) for file_name in file_names]
    return file_paths


def get_tellingen_count():
    return Telling.objects.all().count()

